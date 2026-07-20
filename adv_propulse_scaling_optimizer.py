from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.interpolate import LinearNDInterpolator, NearestNDInterpolator

KNOT_TO_MS = 0.514444
DEFAULT_REF_DIAMETER_M = 0.300
DEFAULT_REF_BLADE_LENGTH_M = 5.00
DEFAULT_BLADES = 3.0


@dataclass(frozen=True)
class ScalingExponents:
    force_diam_exp: float
    moment_diam_exp: float
    power_diam_exp: float


SCALING_MODES = {
    "3d": ScalingExponents(force_diam_exp=2.0, moment_diam_exp=3.0, power_diam_exp=2.0),
    "2d": ScalingExponents(force_diam_exp=1.0, moment_diam_exp=2.0, power_diam_exp=1.0),
}

REYNOLDS_MODES = {
    "diameter": "Iso-Re approché sur V*D",
    "chord_ratio": "Iso-Re sur V*c via c_target/c_ref imposé",
    "sigma": "Iso-Re sur V*c via c ~ sigma*D/Z",
}


class Surrogate3D:
    """Interpolateur robuste (linéaire + nearest en secours) sur (V, lambda, Bmax)."""

    def __init__(self, df: pd.DataFrame, input_cols: Iterable[str], output_cols: Iterable[str]) -> None:
        self.input_cols = list(input_cols)
        self.output_cols = list(output_cols)
        self.df = df.copy()

        pts = df[self.input_cols].to_numpy(dtype=float)
        self._linear: Dict[str, LinearNDInterpolator] = {}
        self._nearest: Dict[str, NearestNDInterpolator] = {}
        for col in self.output_cols:
            vals = df[col].to_numpy(dtype=float)
            self._linear[col] = LinearNDInterpolator(pts, vals, fill_value=np.nan)
            self._nearest[col] = NearestNDInterpolator(pts, vals)

        self.v_min = float(df[self.input_cols[0]].min())
        self.v_max = float(df[self.input_cols[0]].max())
        self.lambda_min = float(df[self.input_cols[1]].min())
        self.lambda_max = float(df[self.input_cols[1]].max())
        self.bmax_min = float(df[self.input_cols[2]].min())
        self.bmax_max = float(df[self.input_cols[2]].max())
        self.lambda_candidates = sorted(df[self.input_cols[1]].dropna().unique().astype(float).tolist())
        self.bmax_candidates = sorted(df[self.input_cols[2]].dropna().unique().astype(float).tolist())
        self.v_candidates = sorted(df[self.input_cols[0]].dropna().unique().astype(float).tolist())
        self.database_orbital_diameter_m = float(df.attrs.get("database_orbital_diameter_m", DEFAULT_REF_DIAMETER_M))
        self.cleaning_report = dict(df.attrs.get("cleaning_report", {}))

    def __call__(self, V_ms: float, lam: float, bmax_deg: float) -> Dict[str, float]:
        p = np.array([[float(V_ms), float(lam), float(bmax_deg)]], dtype=float)
        out: Dict[str, float] = {}
        used_nearest = False
        for col in self.output_cols:
            v = float(self._linear[col](p)[0])
            if np.isnan(v):
                v = float(self._nearest[col](p)[0])
                used_nearest = True
            out[col] = v
        out["_used_nearest_fallback"] = used_nearest
        return out

    def domain_diagnostics(self, V_ms: float, lam: float, bmax_deg: float) -> Dict[str, object]:
        diagnostics = {
            "database_V_min_ms": self.v_min,
            "database_V_max_ms": self.v_max,
            "database_lambda_min": self.lambda_min,
            "database_lambda_max": self.lambda_max,
            "database_Bmax_min_deg": self.bmax_min,
            "database_Bmax_max_deg": self.bmax_max,
            "V_ref_in_domain": self.v_min <= V_ms <= self.v_max,
            "lambda_in_domain": self.lambda_min <= lam <= self.lambda_max,
            "Bmax_in_domain": self.bmax_min <= bmax_deg <= self.bmax_max,
        }
        diagnostics["domain_ok"] = bool(
            diagnostics["V_ref_in_domain"] and diagnostics["lambda_in_domain"] and diagnostics["Bmax_in_domain"]
        )
        diagnostics["domain_warning"] = "; ".join(
            [
                msg
                for cond, msg in [
                    (diagnostics["V_ref_in_domain"], "V_ref hors domaine"),
                    (diagnostics["lambda_in_domain"], "lambda hors domaine"),
                    (diagnostics["Bmax_in_domain"], "Bmax hors domaine"),
                ]
                if not cond
            ]
        )
        return diagnostics




def _normalize_power_of_1000(value: float, lower: float, upper: float) -> float:
    """Ramène une valeur dans une plage plausible en corrigeant les pertes de séparateur décimal x1000."""
    x = float(value)
    if not np.isfinite(x) or x == 0:
        return x
    ax = abs(x)
    sign = -1.0 if x < 0 else 1.0
    for _ in range(4):
        if ax < lower:
            ax *= 1000.0
        elif ax > upper:
            ax /= 1000.0
        else:
            break
    return sign * ax


def infer_database_orbital_diameter(df: pd.DataFrame, fallback_m: float = DEFAULT_REF_DIAMETER_M) -> float:
    """Déduit D_orb de lambda = V/(omega D/2), en corrigeant les x1000 présents dans le classeur."""
    required = {"Ve [m/s]", "Lambda [-]", "omega [rad/s]"}
    if not required.issubset(df.columns):
        return float(fallback_m)
    v = pd.to_numeric(df["Ve [m/s]"], errors="coerce").to_numpy(dtype=float)
    lam = pd.to_numeric(df["Lambda [-]"], errors="coerce").to_numpy(dtype=float)
    omega = pd.to_numeric(df["omega [rad/s]"], errors="coerce").to_numpy(dtype=float)
    with np.errstate(divide="ignore", invalid="ignore"):
        candidates = 2.0 * v / (lam * omega)
    normalized = []
    for x in candidates:
        if not np.isfinite(x) or x <= 0:
            continue
        normalized.append(_normalize_power_of_1000(float(x), 0.02, 20.0))
    normalized = np.asarray([x for x in normalized if 0.02 <= x <= 20.0], dtype=float)
    if normalized.size < 10:
        return float(fallback_m)
    return float(np.median(normalized))


def _repair_v2_scaled_column(data: pd.DataFrame, column: str) -> int:
    """Corrige des x1000 isolés sur une grandeur ~V², à lambda/Bmax fixés."""
    if column not in data.columns:
        return 0
    corrected = 0
    out = data[column].astype(float).copy()
    for _, idx in data.groupby(["Lambda [-]", "Bmax[°]"], dropna=False).groups.items():
        idx = list(idx)
        vals = out.loc[idx].to_numpy(dtype=float)
        speeds = data.loc[idx, "Ve [m/s]"].to_numpy(dtype=float)
        with np.errstate(divide="ignore", invalid="ignore"):
            norm = np.abs(vals) / np.maximum(speeds ** 2, 1e-12)
            logs = np.log10(norm)
        finite = np.isfinite(logs)
        if finite.sum() < 3:
            continue
        center = float(np.median(logs[finite]))
        for j, row_idx in enumerate(idx):
            if not np.isfinite(vals[j]) or vals[j] == 0:
                continue
            factors = np.array([0.001, 1.0, 1000.0])
            cand = np.abs(vals[j] * factors) / max(speeds[j] ** 2, 1e-12)
            dist = np.abs(np.log10(cand) - center)
            factor = float(factors[int(np.argmin(dist))])
            if factor != 1.0:
                out.loc[row_idx] = vals[j] * factor
                corrected += 1
    data[column] = out
    return corrected


def sanitize_summary_data(data: pd.DataFrame, database_diameter_m: float) -> Dict[str, int]:
    """Répare les pertes aléatoires de séparateur décimal du fichier sans modifier DHP, utilisé comme ancre."""
    report: Dict[str, int] = {}
    v = data["Ve [m/s]"].to_numpy(dtype=float)
    lam = data["Lambda [-]"].to_numpy(dtype=float)
    omega = 2.0 * v / (lam * float(database_diameter_m))

    if "omega [rad/s]" in data.columns:
        old = data["omega [rad/s]"].to_numpy(dtype=float)
        report["omega_repaired"] = int(np.sum(~np.isclose(old, omega, rtol=1e-4, atol=1e-6)))
        data["omega [rad/s]"] = omega
    if "omega [rpm]" in data.columns:
        rpm = omega * 60.0 / (2.0 * np.pi)
        old = data["omega [rpm]"].to_numpy(dtype=float)
        report["rpm_repaired"] = int(np.sum(~np.isclose(old, rpm, rtol=1e-4, atol=1e-4)))
        data["omega [rpm]"] = rpm

    # DHP est cohérent et lisse dans la base. On l'utilise pour réparer le couple moyen.
    if "DHP[W]" in data.columns and "Cp_mean[N.m]" in data.columns:
        q = data["DHP[W]"].to_numpy(dtype=float) / omega
        old = data["Cp_mean[N.m]"].to_numpy(dtype=float)
        report["Cp_mean_repaired"] = int(np.sum(~np.isclose(old, q, rtol=1e-4, atol=1e-5)))
        data["Cp_mean[N.m]"] = q

    # Eta est borné et certaines cellules sont exactement x1000.
    if "eta_Cp[%]" in data.columns:
        eta_old = data["eta_Cp[%]"].to_numpy(dtype=float)
        eta_norm = np.array([_normalize_power_of_1000(x, 0.01, 100.0) for x in eta_old])
        report["eta_repaired"] = int(np.sum(~np.isclose(eta_old, eta_norm, rtol=1e-8, atol=1e-10)))
        data["eta_Cp[%]"] = eta_norm

    # Corrige la poussée avec la redondance eta = -T V / DHP.
    if {"eta_Cp[%]", "thrust[N]", "DHP[W]"}.issubset(data.columns):
        eta = data["eta_Cp[%]"].to_numpy(dtype=float)
        raw_t = data["thrust[N]"].to_numpy(dtype=float)
        power = data["DHP[W]"].to_numpy(dtype=float)
        repaired_t = raw_t.copy()
        count = 0
        for i in range(len(data)):
            if not all(np.isfinite(x) for x in [eta[i], raw_t[i], power[i], v[i]]) or power[i] == 0 or v[i] == 0:
                continue
            factors = np.array([0.001, 1.0, 1000.0])
            eta_candidates = -raw_t[i] * factors * v[i] / power[i] * 100.0
            valid = eta_candidates > 0
            errors = np.where(valid, np.abs(np.log(np.maximum(eta_candidates, 1e-12) / max(eta[i], 1e-12))), np.inf)
            factor = float(factors[int(np.argmin(errors))])
            if factor != 1.0:
                repaired_t[i] = raw_t[i] * factor
                count += 1
        data["thrust[N]"] = repaired_t
        data["eta_Cp[%]"] = -repaired_t * v / power * 100.0
        report["thrust_repaired"] = count

    for col in ["sideforce[N]", "Mh max", "Cd_mean[N.m]"]:
        report[f"{col}_repaired"] = _repair_v2_scaled_column(data, col)
    return report


def knots_to_ms(knots: float) -> float:
    return knots * KNOT_TO_MS


def ms_to_knots(ms: float) -> float:
    return ms / KNOT_TO_MS


def _pick_sheet_name(wb, preferred: str) -> str:
    if preferred in wb.sheetnames:
        return preferred
    for candidate in ["Summary", "summary", "Sheet1"]:
        if candidate in wb.sheetnames:
            return candidate
    return wb.sheetnames[0]


def load_summary(path: str | Path, summary_sheet: str = "Summary") -> pd.DataFrame:
    """Charge la feuille de synthèse. Si Summary n'existe pas, prend Sheet1 ou la première feuille."""
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    if not wb.sheetnames:
        raise ValueError("Le classeur ne contient aucune feuille.")
    sheet_name = _pick_sheet_name(wb, summary_sheet)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"La feuille '{sheet_name}' est vide.")

    headers = list(rows[0])
    if not headers:
        raise ValueError(f"La feuille '{sheet_name}' ne contient pas d'en-têtes lisibles.")
    if headers[0] is None:
        headers[0] = "run"
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers)]

    data = pd.DataFrame(rows[1:], columns=headers)
    data = data.dropna(how="all").copy()

    for col in data.columns:
        if col == "run":
            continue
        data[col] = pd.to_numeric(data[col], errors="coerce")

    required = ["Ve [m/s]", "Lambda [-]", "Bmax[°]"]
    missing = [c for c in required if c not in data.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes dans '{sheet_name}': {missing}")

    data = data.dropna(subset=required).reset_index(drop=True)

    database_diameter_m = infer_database_orbital_diameter(data)
    cleaning_report = sanitize_summary_data(data, database_diameter_m)

    if "thrust[N]" in data.columns:
        data["thrust_propulsive[N]"] = -data["thrust[N]"]
    if "sideforce[N]" in data.columns:
        data["sideforce_abs[N]"] = data["sideforce[N]"].abs()

    data.attrs["sheet_name"] = sheet_name
    data.attrs["database_orbital_diameter_m"] = float(database_diameter_m)
    data.attrs["cleaning_report"] = cleaning_report
    return data


def build_surrogate(df: pd.DataFrame) -> Surrogate3D:
    needed = [
        "eta_Cp[%]",
        "thrust[N]",
        "thrust_propulsive[N]",
        "sideforce[N]",
        "sideforce_abs[N]",
        "Mh max",
        "DHP[W]",
        "Cp_mean[N.m]",
        "Cd_mean[N.m]",
        "kth[-]",
        "kqh[-]",
        "ktvoith[-]",
        "kqvoith[-]",
        "ks[-]",
        "kd[-]",
    ]
    output_cols = [c for c in needed if c in df.columns]
    return Surrogate3D(df, ["Ve [m/s]", "Lambda [-]", "Bmax[°]"], output_cols)


def chord_ratio_from_sigma(
    D_target_m: float,
    D_ref_m: float,
    sigma_target: float,
    sigma_ref: float,
    blades_target: float = DEFAULT_BLADES,
    blades_ref: float = DEFAULT_BLADES,
) -> float:
    if sigma_ref == 0 or blades_target == 0 or blades_ref == 0:
        raise ValueError("sigma_ref, blades_target et blades_ref doivent être non nuls.")
    return (sigma_target * D_target_m / blades_target) / (sigma_ref * D_ref_m / blades_ref)


def equivalent_reference_speed(
    V_target_ms: float,
    D_target_m: float,
    D_ref_m: float = DEFAULT_REF_DIAMETER_M,
    iso_reynolds: bool = True,
    reynolds_mode: str = "diameter",
    chord_ratio_target_to_ref: Optional[float] = None,
    sigma_target: Optional[float] = None,
    sigma_ref: Optional[float] = None,
    blades_target: float = DEFAULT_BLADES,
    blades_ref: float = DEFAULT_BLADES,
) -> Tuple[float, float]:
    if not iso_reynolds:
        return V_target_ms, 1.0

    if reynolds_mode == "diameter":
        ratio = D_target_m / D_ref_m
    elif reynolds_mode == "chord_ratio":
        if chord_ratio_target_to_ref is None:
            raise ValueError("Le mode chord_ratio nécessite chord_ratio_target_to_ref.")
        ratio = float(chord_ratio_target_to_ref)
    elif reynolds_mode == "sigma":
        if sigma_target is None or sigma_ref is None:
            raise ValueError("Le mode sigma nécessite sigma_target et sigma_ref.")
        ratio = chord_ratio_from_sigma(
            D_target_m=D_target_m,
            D_ref_m=D_ref_m,
            sigma_target=float(sigma_target),
            sigma_ref=float(sigma_ref),
            blades_target=float(blades_target),
            blades_ref=float(blades_ref),
        )
    else:
        raise ValueError(f"reynolds_mode doit être parmi {list(REYNOLDS_MODES)}")

    return V_target_ms * ratio, ratio


def _blade_length_ratio(
    scaling_mode: str,
    D_target_m: float,
    D_ref_m: float,
    target_blade_length_m: Optional[float],
    ref_blade_length_m: Optional[float],
) -> float:
    if scaling_mode == "3d":
        return D_target_m / D_ref_m
    if scaling_mode == "2d":
        if target_blade_length_m is None or ref_blade_length_m is None:
            return 1.0
        if ref_blade_length_m == 0:
            raise ValueError("La longueur de pale de référence doit être non nulle.")
        return float(target_blade_length_m) / float(ref_blade_length_m)
    raise ValueError(f"scaling_mode doit être parmi {list(SCALING_MODES)}")


def rescale_outputs(
    base: Dict[str, float],
    V_target_ms: float,
    V_ref_ms: float,
    D_target_m: float,
    D_ref_m: float,
    scaling_mode: str = "3d",
    target_blade_length_m: Optional[float] = None,
    ref_blade_length_m: Optional[float] = None,
) -> Dict[str, float]:
    if scaling_mode not in SCALING_MODES:
        raise ValueError(f"scaling_mode doit être parmi {list(SCALING_MODES)}")
    if V_ref_ms == 0:
        raise ValueError("V_ref_ms ne peut pas être nul.")

    exps = SCALING_MODES[scaling_mode]
    out = dict(base)
    v_ratio = float(V_target_ms / V_ref_ms)
    d_ratio = float(D_target_m / D_ref_m)
    blade_length_ratio = float(
        _blade_length_ratio(
            scaling_mode=scaling_mode,
            D_target_m=D_target_m,
            D_ref_m=D_ref_m,
            target_blade_length_m=target_blade_length_m,
            ref_blade_length_m=ref_blade_length_m,
        )
    )
    h_exp = 1.0 if scaling_mode == "2d" else 0.0

    def scale_value(name: str, v_exp: float, d_exp: float, h_exp_local: float) -> float:
        if name not in out or out[name] is None:
            return np.nan
        factor = (v_ratio ** v_exp) * (d_ratio ** d_exp) * (blade_length_ratio ** h_exp_local)
        out[name] = float(out[name]) * factor
        return float(factor)

    force_factor = scale_value("thrust[N]", 2.0, exps.force_diam_exp, h_exp)
    scale_value("thrust_propulsive[N]", 2.0, exps.force_diam_exp, h_exp)
    scale_value("sideforce[N]", 2.0, exps.force_diam_exp, h_exp)
    scale_value("sideforce_abs[N]", 2.0, exps.force_diam_exp, h_exp)

    moment_factor = scale_value("Mh max", 2.0, exps.moment_diam_exp, h_exp)
    scale_value("Cp_mean[N.m]", 2.0, exps.moment_diam_exp, h_exp)
    scale_value("Cd_mean[N.m]", 2.0, exps.moment_diam_exp, h_exp)

    power_factor = scale_value("DHP[W]", 3.0, exps.power_diam_exp, h_exp)

    if "thrust[N]" in out:
        out["thrust_propulsive[N]"] = -float(out["thrust[N]"])
    if "sideforce[N]" in out:
        out["sideforce_abs[N]"] = abs(float(out["sideforce[N]"]))

    out.update(
        {
            "V_ratio_target_over_database": v_ratio,
            "D_ratio_target_over_reference": d_ratio,
            "blade_length_ratio_used": blade_length_ratio,
            "force_scale_factor": float(force_factor) if np.isfinite(force_factor) else np.nan,
            "moment_scale_factor": float(moment_factor) if np.isfinite(moment_factor) else np.nan,
            "power_scale_factor": float(power_factor) if np.isfinite(power_factor) else np.nan,
        }
    )
    return out


def evaluate_target_case(
    surrogate: Surrogate3D,
    V_target_ms: float,
    lam: float,
    bmax_deg: float,
    D_target_m: float,
    D_ref_m: float = DEFAULT_REF_DIAMETER_M,
    iso_reynolds: bool = True,
    scaling_mode: str = "3d",
    reynolds_mode: str = "diameter",
    chord_ratio_target_to_ref: Optional[float] = None,
    sigma_target: Optional[float] = None,
    sigma_ref: Optional[float] = None,
    blades_target: float = DEFAULT_BLADES,
    blades_ref: float = DEFAULT_BLADES,
    target_blade_length_m: Optional[float] = None,
    ref_blade_length_m: Optional[float] = None,
) -> Dict[str, float]:
    inferred_ref = float(getattr(surrogate, "database_orbital_diameter_m", D_ref_m))
    if not np.isclose(float(D_ref_m), inferred_ref, rtol=0.02, atol=1e-6):
        raise ValueError(
            f"Diametre de reference incoherent avec la base: {D_ref_m:.6g} m fourni, "
            f"{inferred_ref:.6g} m deduit de V, lambda et omega. "
            "Le diametre de base ne doit pas etre remplace par une dimension de la machine reelle."
        )

    V_ref_ms, chord_ratio_used = equivalent_reference_speed(
        V_target_ms=V_target_ms,
        D_target_m=D_target_m,
        D_ref_m=D_ref_m,
        iso_reynolds=iso_reynolds,
        reynolds_mode=reynolds_mode,
        chord_ratio_target_to_ref=chord_ratio_target_to_ref,
        sigma_target=sigma_target,
        sigma_ref=sigma_ref,
        blades_target=blades_target,
        blades_ref=blades_ref,
    )
    diagnostics = surrogate.domain_diagnostics(V_ref_ms, lam, bmax_deg)
    base = surrogate(V_ref_ms, lam, bmax_deg)
    used_nearest_fallback = bool(base.pop("_used_nearest_fallback", False))

    scaled = rescale_outputs(
        base=base,
        V_target_ms=V_target_ms,
        V_ref_ms=V_ref_ms,
        D_target_m=D_target_m,
        D_ref_m=D_ref_m,
        scaling_mode=scaling_mode,
        target_blade_length_m=target_blade_length_m,
        ref_blade_length_m=ref_blade_length_m,
    )

    omega = 2.0 * V_target_ms / (lam * D_target_m)
    scaled.update(
        {
            "V_target_ms": float(V_target_ms),
            "V_target_kn": float(ms_to_knots(V_target_ms)),
            "lambda": float(lam),
            "Bmax_deg": float(bmax_deg),
            "D_target_m": float(D_target_m),
            "D_ref_m": float(D_ref_m),
            "V_reference_in_database_ms": float(V_ref_ms),
            "V_reference_in_database_kn": float(ms_to_knots(V_ref_ms)),
            "omega_target_rad_s": float(omega),
            "omega_target_rpm": float(omega * 60.0 / (2.0 * np.pi)),
            "iso_reynolds": bool(iso_reynolds),
            "reynolds_mode": reynolds_mode,
            "chord_ratio_target_to_ref_used": float(chord_ratio_used),
            "sigma_target": np.nan if sigma_target is None else float(sigma_target),
            "sigma_ref": np.nan if sigma_ref is None else float(sigma_ref),
            "blades_target": float(blades_target),
            "blades_ref": float(blades_ref),
            "target_blade_length_m": np.nan if target_blade_length_m is None else float(target_blade_length_m),
            "ref_blade_length_m": np.nan if ref_blade_length_m is None else float(ref_blade_length_m),
            "scaling_mode": scaling_mode,
            "used_nearest_fallback": used_nearest_fallback,
        }
    )
    scaled.update(diagnostics)
    return scaled


def _apply_constraints(
    table: pd.DataFrame,
    min_thrust_propulsive_N: Optional[float] = None,
    max_Mh_Nm: Optional[float] = None,
    max_DHP_W: Optional[float] = None,
) -> pd.DataFrame:
    table = table.copy()
    statuses = []
    feasible_flags = []
    for _, row in table.iterrows():
        status = "ok"
        if min_thrust_propulsive_N is not None and row.get("thrust_propulsive[N]", -np.inf) < min_thrust_propulsive_N:
            status = "below_min_thrust"
        if max_Mh_Nm is not None and row.get("Mh max", np.inf) > max_Mh_Nm:
            status = "above_max_Mh"
        if max_DHP_W is not None and row.get("DHP[W]", np.inf) > max_DHP_W:
            status = "above_max_DHP"
        feasible_flags.append(status == "ok")
        statuses.append(status)
    table["feasible"] = feasible_flags
    table["status"] = statuses
    if "DHP[W]" in table.columns and "thrust_propulsive[N]" in table.columns:
        table["thrust_per_power"] = table["thrust_propulsive[N]"] / table["DHP[W]"].clip(lower=1e-9)
    return table


def _best_index(feasible: pd.DataFrame, objective: str) -> int:
    if objective == "max_eta":
        return int(feasible["eta_Cp[%]"].idxmax())
    if objective == "min_dhp":
        return int(feasible["DHP[W]"].idxmin())
    if objective == "max_thrust":
        return int(feasible["thrust_propulsive[N]"].idxmax())
    if objective == "max_thrust_per_power":
        return int(feasible["thrust_per_power"].idxmax())
    raise ValueError("objective doit être parmi: max_eta, min_dhp, max_thrust, max_thrust_per_power")


def optimize_operating_point(
    surrogate: Surrogate3D,
    V_target_ms: float,
    D_target_m: float,
    lambda_candidates: Iterable[float],
    bmax_candidates: Iterable[float],
    D_ref_m: float = DEFAULT_REF_DIAMETER_M,
    iso_reynolds: bool = True,
    scaling_mode: str = "3d",
    objective: str = "max_eta",
    min_thrust_propulsive_N: Optional[float] = None,
    max_Mh_Nm: Optional[float] = None,
    max_DHP_W: Optional[float] = None,
    reynolds_mode: str = "diameter",
    chord_ratio_target_to_ref: Optional[float] = None,
    sigma_target: Optional[float] = None,
    sigma_ref: Optional[float] = None,
    blades_target: float = DEFAULT_BLADES,
    blades_ref: float = DEFAULT_BLADES,
    target_blade_length_m: Optional[float] = None,
    ref_blade_length_m: Optional[float] = None,
) -> Tuple[Dict[str, float], pd.DataFrame]:
    rows = []
    for lam in lambda_candidates:
        for bmax in bmax_candidates:
            rows.append(
                evaluate_target_case(
                    surrogate=surrogate,
                    V_target_ms=V_target_ms,
                    lam=float(lam),
                    bmax_deg=float(bmax),
                    D_target_m=D_target_m,
                    D_ref_m=D_ref_m,
                    iso_reynolds=iso_reynolds,
                    scaling_mode=scaling_mode,
                    reynolds_mode=reynolds_mode,
                    chord_ratio_target_to_ref=chord_ratio_target_to_ref,
                    sigma_target=sigma_target,
                    sigma_ref=sigma_ref,
                    blades_target=blades_target,
                    blades_ref=blades_ref,
                    target_blade_length_m=target_blade_length_m,
                    ref_blade_length_m=ref_blade_length_m,
                )
            )

    table = _apply_constraints(
        pd.DataFrame(rows),
        min_thrust_propulsive_N=min_thrust_propulsive_N,
        max_Mh_Nm=max_Mh_Nm,
        max_DHP_W=max_DHP_W,
    )
    feasible = table[table["feasible"]].copy()
    if feasible.empty:
        raise RuntimeError("Aucun point faisable avec les contraintes données.")

    idx = _best_index(feasible, objective)
    best = feasible.loc[idx].to_dict()
    return best, table.sort_values(["lambda", "Bmax_deg"]).reset_index(drop=True)


def suggest_raw_bmax_envelope(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    idx = work.groupby("Lambda [-]")["eta_Cp[%]"].idxmax()
    law = work.loc[idx, ["Ve [m/s]", "Lambda [-]", "Bmax[°]", "eta_Cp[%]"]].sort_values("Lambda [-]")
    law = law.rename(
        columns={
            "Ve [m/s]": "V_base_ms",
            "Lambda [-]": "lambda",
            "Bmax[°]": "Bmax_opt_deg",
            "eta_Cp[%]": "eta_opt_pct",
        }
    )
    return law.reset_index(drop=True)


def suggest_bmax_law_for_target(
    surrogate: Surrogate3D,
    V_target_ms: float,
    D_target_m: float,
    lambda_candidates: Iterable[float],
    bmax_candidates: Iterable[float],
    D_ref_m: float = DEFAULT_REF_DIAMETER_M,
    iso_reynolds: bool = True,
    scaling_mode: str = "3d",
    objective: str = "max_eta",
    min_thrust_propulsive_N: Optional[float] = None,
    max_Mh_Nm: Optional[float] = None,
    max_DHP_W: Optional[float] = None,
    reynolds_mode: str = "diameter",
    chord_ratio_target_to_ref: Optional[float] = None,
    sigma_target: Optional[float] = None,
    sigma_ref: Optional[float] = None,
    blades_target: float = DEFAULT_BLADES,
    blades_ref: float = DEFAULT_BLADES,
    target_blade_length_m: Optional[float] = None,
    ref_blade_length_m: Optional[float] = None,
) -> pd.DataFrame:
    rows = []
    for lam in lambda_candidates:
        try:
            best, _ = optimize_operating_point(
                surrogate=surrogate,
                V_target_ms=V_target_ms,
                D_target_m=D_target_m,
                lambda_candidates=[lam],
                bmax_candidates=bmax_candidates,
                D_ref_m=D_ref_m,
                iso_reynolds=iso_reynolds,
                scaling_mode=scaling_mode,
                objective=objective,
                min_thrust_propulsive_N=min_thrust_propulsive_N,
                max_Mh_Nm=max_Mh_Nm,
                max_DHP_W=max_DHP_W,
                reynolds_mode=reynolds_mode,
                chord_ratio_target_to_ref=chord_ratio_target_to_ref,
                sigma_target=sigma_target,
                sigma_ref=sigma_ref,
                blades_target=blades_target,
                blades_ref=blades_ref,
                target_blade_length_m=target_blade_length_m,
                ref_blade_length_m=ref_blade_length_m,
            )
            rows.append(
                {
                    "lambda": best["lambda"],
                    "feasible": True,
                    "status": "ok",
                    "Bmax_opt_deg": best["Bmax_deg"],
                    "eta_opt_pct": best.get("eta_Cp[%]"),
                    "thrust_propulsive_N": best.get("thrust_propulsive[N]"),
                    "sideforce_abs_N": best.get("sideforce_abs[N]"),
                    "Mh_max_Nm": best.get("Mh max"),
                    "DHP_W": best.get("DHP[W]"),
                    "omega_rpm": best.get("omega_target_rpm"),
                    "V_reference_in_database_ms": best.get("V_reference_in_database_ms"),
                    "domain_warning": best.get("domain_warning"),
                }
            )
        except RuntimeError:
            rows.append(
                {
                    "lambda": float(lam),
                    "feasible": False,
                    "status": "no_feasible_point",
                    "Bmax_opt_deg": np.nan,
                    "eta_opt_pct": np.nan,
                    "thrust_propulsive_N": np.nan,
                    "sideforce_abs_N": np.nan,
                    "Mh_max_Nm": np.nan,
                    "DHP_W": np.nan,
                    "omega_rpm": np.nan,
                    "V_reference_in_database_ms": np.nan,
                    "domain_warning": "",
                }
            )
    return pd.DataFrame(rows).sort_values("lambda").reset_index(drop=True)


def format_best_point(best: Dict[str, float]) -> str:
    keys = [
        "lambda",
        "Bmax_deg",
        "eta_Cp[%]",
        "thrust_propulsive[N]",
        "sideforce_abs[N]",
        "Mh max",
        "DHP[W]",
        "omega_target_rad_s",
        "omega_target_rpm",
        "V_reference_in_database_ms",
        "blade_length_ratio_used",
        "force_scale_factor",
        "moment_scale_factor",
        "power_scale_factor",
        "reynolds_mode",
        "chord_ratio_target_to_ref_used",
        "domain_warning",
        "used_nearest_fallback",
    ]
    return "\n".join(f"{k}: {best[k]}" for k in keys if k in best)


def _add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("xlsx", help="Chemin du fichier xlsx")
    parser.add_argument("--diameter-mm", type=float, required=True, help="Diamètre orbital cible en mm")
    parser.add_argument("--speed-kn", type=float, required=True, help="Vitesse cible en noeuds")
    parser.add_argument("--reference-diameter-mm", type=float, default=300.0)
    parser.add_argument("--scaling-mode", choices=list(SCALING_MODES), default="3d")
    parser.add_argument("--no-iso-re", action="store_true", help="Utilise le même V dans la base au lieu du cas iso-Re")
    parser.add_argument("--reynolds-mode", choices=list(REYNOLDS_MODES), default="diameter")
    parser.add_argument("--chord-ratio-target-to-ref", type=float, default=None)
    parser.add_argument("--sigma-target", type=float, default=None)
    parser.add_argument("--sigma-ref", type=float, default=None)
    parser.add_argument("--blades-target", type=float, default=DEFAULT_BLADES)
    parser.add_argument("--blades-ref", type=float, default=DEFAULT_BLADES)
    parser.add_argument("--target-blade-length-m", type=float, default=None)
    parser.add_argument("--ref-blade-length-m", type=float, default=DEFAULT_REF_BLADE_LENGTH_M)


def main() -> None:
    parser = argparse.ArgumentParser(description="Mise à l'échelle et optimisation d'un ADV-propulse à partir d'une base de résultats CFD.")
    _add_common_args(parser)
    parser.add_argument("--lambda", dest="lam", type=float, default=None, help="Lambda imposé")
    parser.add_argument("--bmax", type=float, default=None, help="Bmax imposé en degrés")
    parser.add_argument("--optimize", action="store_true", help="Optimise lambda/Bmax au lieu d'évaluer un point imposé")
    parser.add_argument("--objective", choices=["max_eta", "min_dhp", "max_thrust", "max_thrust_per_power"], default="max_eta")
    parser.add_argument("--min-thrust-N", type=float, default=None)
    parser.add_argument("--max-Mh-Nm", type=float, default=None)
    parser.add_argument("--max-DHP-W", type=float, default=None)
    args = parser.parse_args()

    df = load_summary(args.xlsx)
    surrogate = build_surrogate(df)

    V_target_ms = knots_to_ms(args.speed_kn)
    D_target_m = args.diameter_mm / 1000.0
    D_ref_m = args.reference_diameter_mm / 1000.0
    iso_reynolds = not args.no_iso_re

    if args.optimize:
        lambda_candidates = surrogate.lambda_candidates
        bmax_candidates = surrogate.bmax_candidates
        best, table = optimize_operating_point(
            surrogate=surrogate,
            V_target_ms=V_target_ms,
            D_target_m=D_target_m,
            lambda_candidates=lambda_candidates,
            bmax_candidates=bmax_candidates,
            D_ref_m=D_ref_m,
            iso_reynolds=iso_reynolds,
            scaling_mode=args.scaling_mode,
            objective=args.objective,
            min_thrust_propulsive_N=args.min_thrust_N,
            max_Mh_Nm=args.max_Mh_Nm,
            max_DHP_W=args.max_DHP_W,
            reynolds_mode=args.reynolds_mode,
            chord_ratio_target_to_ref=args.chord_ratio_target_to_ref,
            sigma_target=args.sigma_target,
            sigma_ref=args.sigma_ref,
            blades_target=args.blades_target,
            blades_ref=args.blades_ref,
            target_blade_length_m=args.target_blade_length_m,
            ref_blade_length_m=args.ref_blade_length_m,
        )
        print("=== Meilleur point ===")
        print(format_best_point(best))
        print("\n=== Loi discrète Bmax_opt(lambda) cohérente avec le cas cible ===")
        law = suggest_bmax_law_for_target(
            surrogate=surrogate,
            V_target_ms=V_target_ms,
            D_target_m=D_target_m,
            lambda_candidates=lambda_candidates,
            bmax_candidates=bmax_candidates,
            D_ref_m=D_ref_m,
            iso_reynolds=iso_reynolds,
            scaling_mode=args.scaling_mode,
            objective=args.objective,
            min_thrust_propulsive_N=args.min_thrust_N,
            max_Mh_Nm=args.max_Mh_Nm,
            max_DHP_W=args.max_DHP_W,
            reynolds_mode=args.reynolds_mode,
            chord_ratio_target_to_ref=args.chord_ratio_target_to_ref,
            sigma_target=args.sigma_target,
            sigma_ref=args.sigma_ref,
            blades_target=args.blades_target,
            blades_ref=args.blades_ref,
            target_blade_length_m=args.target_blade_length_m,
            ref_blade_length_m=args.ref_blade_length_m,
        )
        print(law.to_string(index=False))
        print("\n=== Enveloppe brute de la base CFD (indépendante du cas cible) ===")
        print(suggest_raw_bmax_envelope(df).to_string(index=False))
        print("\n=== Top 10 faisables ===")
        sort_col = {
            "max_eta": "eta_Cp[%]",
            "min_dhp": "DHP[W]",
            "max_thrust": "thrust_propulsive[N]",
            "max_thrust_per_power": "thrust_per_power",
        }[args.objective]
        ascending = args.objective == "min_dhp"
        cols = [
            c
            for c in [
                "lambda",
                "Bmax_deg",
                "eta_Cp[%]",
                "thrust_propulsive[N]",
                "sideforce_abs[N]",
                "Mh max",
                "DHP[W]",
                "thrust_per_power",
                "domain_warning",
            ]
            if c in table.columns
        ]
        print(table[table["feasible"]][cols].sort_values(sort_col, ascending=ascending).head(10).to_string(index=False))
        return

    if args.lam is None or args.bmax is None:
        parser.error("En mode évaluation d'un point, il faut fournir --lambda et --bmax.")

    point = evaluate_target_case(
        surrogate=surrogate,
        V_target_ms=V_target_ms,
        lam=args.lam,
        bmax_deg=args.bmax,
        D_target_m=D_target_m,
        D_ref_m=D_ref_m,
        iso_reynolds=iso_reynolds,
        scaling_mode=args.scaling_mode,
        reynolds_mode=args.reynolds_mode,
        chord_ratio_target_to_ref=args.chord_ratio_target_to_ref,
        sigma_target=args.sigma_target,
        sigma_ref=args.sigma_ref,
        blades_target=args.blades_target,
        blades_ref=args.blades_ref,
        target_blade_length_m=args.target_blade_length_m,
        ref_blade_length_m=args.ref_blade_length_m,
    )
    print("=== Evaluation d'un point ===")
    print(format_best_point(point))


if __name__ == "__main__":
    main()
